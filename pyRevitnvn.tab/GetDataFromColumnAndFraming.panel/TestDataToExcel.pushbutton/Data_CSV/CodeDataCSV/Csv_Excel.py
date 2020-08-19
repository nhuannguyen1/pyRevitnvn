from openpyxl.styles import Alignment
from ReturnDataAllRowByIndexpath import ReturnDataAllRowByIndexpath
from Path_Connect_Excel import Right_Genneral_All_path,Left_Genneral_All_path,DataExcel,Config_Setting_Path

ValueGeneral = [int(i) for i in ReturnDataAllRowByIndexpath(Config_Setting_Path,12)]
#remove Column move 
#Columnmove = ReturnDataAllRowByIndexpath(Config_Setting_Path,32)
Columnmove = [int(i) for i in ReturnDataAllRowByIndexpath(Config_Setting_Path,16)]

LocationCellForMoveColumn = ReturnDataAllRowByIndexpath(Config_Setting_Path,17)

ArrRangMax = [* range(max([int(i) for i in ReturnDataAllRowByIndexpath(Config_Setting_Path,11)] + Columnmove + [int(i) for i in  ReturnDataAllRowByIndexpath(Config_Setting_Path,0)]) + 1)]

ValueChangeLeft_Right  = set(ArrRangMax).difference(set(ValueGeneral + Columnmove))

IndexChangeRemoveColumn = set([int(i) for i in  ReturnDataAllRowByIndexpath(Config_Setting_Path,0)]).difference(set(Columnmove))

GenneralColumnNotChange =  [int(i) for i in ReturnDataAllRowByIndexpath(Config_Setting_Path,19)]

GeneralConcernRaffter = [int(i) for i in ReturnDataAllRowByIndexpath(Config_Setting_Path,20)]

Genneral_Select = [int(i) for i in ReturnDataAllRowByIndexpath(Config_Setting_Path,21)]

LocationOfRowLeft = [int(i) for i in ReturnDataAllRowByIndexpath(Config_Setting_Path,23)]

LocationOfRowRight = [int(i) for i in ReturnDataAllRowByIndexpath(Config_Setting_Path,24)]

ExcelCellForMoveColumnRight = ReturnDataAllRowByIndexpath(Config_Setting_Path,25)

LocationOfPurlin =  [int(i) for i in ReturnDataAllRowByIndexpath(Config_Setting_Path,26)]

TitleLocationMoveColumn = ReturnDataAllRowByIndexpath(Config_Setting_Path,27)

Index_Path_From_CSV  = ReturnDataAllRowByIndexpath(Config_Setting_Path,28)

startrow = ReturnDataAllRowByIndexpath(Config_Setting_Path,30)

def HandlingDataSTr (Element):
    Element1 = Element.replace(":",",")
    Element2= Element1.replace("(","")
    Element3= Element2.replace(")","")
    return eval(Element3)
def GetIndexOfNotChange(IndexChange,IndexChangeTotal):
    IndexArr =  [IndexChangeTotal.index(indexc) for indexc in IndexChange]
    return  IndexArr
def AligntText(sheet):
    rows = range(1, 500)
    columns = range(1, 44)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
def Duplicate_Row_Dataframe(df):
    for cols in df.columns:
        try:
            value = df.loc[1, cols]
            if value == None:
                value = ""       
        except:
            pass
        df[cols].fillna(value, inplace = True) 
    return df
def WriteMoveColumn (pd,worksheet,path,LocationMoveColumn,Columnmove):
    for cell,index in zip(LocationMoveColumn,Columnmove):
        cellstr = HandlingDataSTr(cell)
        df = pd.read_csv(path, delimiter=',',index_col = None ,usecols = [index],nrows= 1)
        worksheet.cell(row = cellstr[0], column = cellstr[1]).value = df.iat[0,0]

def Write_Path_ToExcel (worksheet,path):
    Location_Of_Cell_Path = ReturnDataAllRowByIndexpath(Config_Setting_Path,29)
    LocCell = [HandlingDataSTr(loc) for loc in Location_Of_Cell_Path]
    if path == Left_Genneral_All_path:
        LocCell_C =  LocCell[0]
    else:
       LocCell_C =  LocCell[1]
    worksheet.cell(row = LocCell_C[0], column = LocCell_C[1]).value = "Path CSV"
    worksheet.cell(row = LocCell_C[0], column = (int(LocCell_C[1]) + 1)).value = path
def Write_Path_ToCSV (worksheet,path):
    Location_Of_Cell_Path = ReturnDataAllRowByIndexpath(Config_Setting_Path,29)
    LocCell = [HandlingDataSTr(loc) for loc in Location_Of_Cell_Path]
    if path == Left_Genneral_All_path:
        LocCell_C =  LocCell[0]
    else:
       LocCell_C =  LocCell[1]
    return LocCell_C
import pandas as pd
from openpyxl import load_workbook
from Path_Connect_Excel import Right_Genneral_All_path,Left_Genneral_All_path,DataExcel,Config_Setting_Path
from Csv_Excel import HandlingDataSTr,GetIndexOfNotChange,AligntText,ValueGeneral,Columnmove,LocationCellForMoveColumn,\
    ValueChangeLeft_Right,IndexChangeRemoveColumn,Genneral_Select,GenneralColumnNotChange,GeneralConcernRaffter,\
        TitleLocationMoveColumn,Duplicate_Row_Dataframe,startrow,LocationOfRowLeft,LocationOfRowRight,LocationOfPurlin,Index_Path_From_CSV,Write_Path_ToCSV
from Create_Dict_Between_Csv_And_File import Update_Dict_Joint
from ReturnDataAllRowByIndexpath import ReturnDataAllRowByIndexpath
def Return_cell_obj_Arr (sheet,minrow,mincol, maxcol, maxrow):
    ArrRafterTT = []
    for row in sheet.iter_cols (min_row = minrow,min_col = mincol, max_col = maxcol, max_row = maxrow):
        ArrRafter = [cell.value for cell in row]
        ArrRafterTT.append(ArrRafter)
    return ArrRafterTT
def ReturnCountRafter (kk,sheet):
        last_row = 0
        while True:
            cell_value = sheet.cell(row = kk, column = 1).value
            if cell_value is not None:              
                last_row += 1
                kk = kk + 1
            else:
                break
        return last_row
def CreateFileCSV():
    book = load_workbook("new_big_file.xlsx")
    #Wirite csv 
    sheet = book.get_sheet_by_name('General Member')
    #Create dict Genneral value 
    cell_obj_Arr = Return_cell_obj_Arr(sheet,int(startrow[0]) + 1,1,len(ValueGeneral),int(startrow[0]) + 2)
    #cell_obj_Arr = [sheet.cell(row = 4, column = (i + 1)).value for i,ValueIndexGenneral in enumerate(ValueGeneral)]
    Dict_Value_Genneral = dict(zip(ValueGeneral,cell_obj_Arr))
    paths = [Left_Genneral_All_path,Right_Genneral_All_path]
    for path in paths:
        if path == Left_Genneral_All_path:
            LocationOfRow = LocationOfRowLeft
            Pathdt = 'Left.csv'
            Cell_Address_Arr = Write_Path_ToCSV(sheet,path)
            Cell_Address = Cell_Address_Arr
        else:
            LocationOfRow = LocationOfRowRight
            Pathdt = 'Right.csv'
            Cell_Address_Arr = Write_Path_ToCSV(sheet,path)
            Cell_Address = Cell_Address_Arr
        #create dict for Genneral select:
        cell_obj_Arr = Return_cell_obj_Arr(sheet,int(LocationOfRow[0]) + 1,1,len(Genneral_Select),int(LocationOfRow[0]) + 2)
        #cell_obj_Arr = [sheet.cell(row = 12, column = (i + 1)).value for i,ValueIndexGenneral in enumerate(Genneral_Select)]
        Dict_Value_Genneral_Select = dict(zip(Genneral_Select,cell_obj_Arr))
        #create dict for column Selected:
        cell_obj_Arr = Return_cell_obj_Arr(sheet,int(LocationOfRow[1]) + 1,1,len(GenneralColumnNotChange),int(LocationOfRow[1]) + 2)
        #cell_obj_Arr = [sheet.cell(row = 22, column = (i + 1)).value for i,ValueIndexGenneral in enumerate(GenneralColumnNotChange)]
        Dict_Column_Select = dict(zip(GenneralColumnNotChange,cell_obj_Arr))
        #create dict for Rafter Selected:
        count = ReturnCountRafter(int(LocationOfRow[2]) + 1,sheet)
        ArrRafterTT = Return_cell_obj_Arr(sheet,int(LocationOfRow[2]) + 1,1,6,int(LocationOfRow[2]) + count)
        Dict_Rafter_Select = dict(zip(GeneralConcernRaffter,ArrRafterTT))
        #create purlin list
        #print (int(LocationOfRow[3]) + 1,1,len(LocationOfPurlin),int(LocationOfRow[3]) + 3)
        Arr_Purlin = Return_cell_obj_Arr(sheet,int(LocationOfRow[3]) + 1,1,len(LocationOfPurlin),int(LocationOfRow[3]) + 2)
        Dict_Purlin = dict(zip(LocationOfPurlin,Arr_Purlin))
        #write path to csv from excel 
        Path_Value = sheet.cell(row = Cell_Address[0], column = int(Cell_Address[1]) + 1).value
        Path_Title = sheet.cell(row = Cell_Address[0], column = Cell_Address[1]).value
        Path_Arr = [Path_Title,Path_Value]
        Dict_Path_Arr = {int(Index_Path_From_CSV[0]):Path_Arr}
        #Dict_Purlin = dict(zip(Index_Path_From_CSV,Path_Value))
        #create dict for move column 
        cell_obj_Arr = [sheet.cell(row =  HandlingDataSTr(cell)[0], column = HandlingDataSTr(cell)[1]).value for cell in LocationCellForMoveColumn]
        Title_Arr = TitleLocationMoveColumn
        cell_obj_Arr = list(zip (Title_Arr,cell_obj_Arr))
        Dict_Column_Move = dict(zip(Columnmove,cell_obj_Arr))
        #update Dicts to a gennenral Dict 
        Genenral_Dict = {}
        ListDist = [Dict_Column_Move,Dict_Rafter_Select,Dict_Column_Select,Dict_Value_Genneral,Dict_Value_Genneral_Select,Dict_Purlin,Dict_Path_Arr]
        Genenral_Dict_Sorted = Update_Dict_Joint(*ListDist)
        #df = pd.DataFrame(dict(Genenral_Dict_Sorted))
        df = pd.DataFrame(dict([ (k,pd.Series(v)) for k,v in Genenral_Dict_Sorted.items()]))
        df_ed = Duplicate_Row_Dataframe(df)
        df_ed.to_csv(Pathdt,index=False,header= False)
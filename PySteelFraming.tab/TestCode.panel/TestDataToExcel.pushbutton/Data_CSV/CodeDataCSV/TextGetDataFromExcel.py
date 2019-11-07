import pandas as pd
from pandas import ExcelWriter,ExcelFile
import numpy as np
from openpyxl import load_workbook
from ReturnDataAllRowByIndexpath import ReturnDataAllRowByIndexpath
from Csv_Excel import HandlingDataSTr,GetIndexOfNotChange,AligntText,ValueGeneral,Columnmove,LocationCellForMoveColumn,ValueChangeLeft_Right,\
    IndexChangeRemoveColumn,GenneralColumnNotChange,GeneralConcernRaffter,Genneral_Select,LocationOfRowLeft,LocationOfRowRight,ExcelCellForMoveColumnRight
from Path_Connect_Excel import Right_Genneral_All_path,Left_Genneral_All_path,DataExcel,Config_Setting_Path
#from Csv_Connect_Data import ReturnDataAllRowByIndexpath
book = load_workbook(DataExcel)
print ("LocationCellForMoveColumn",LocationCellForMoveColumn)
def CreateFileExcel():
    writer = ExcelWriter(DataExcel,engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    #ArrMaxValue = IndexGeneral + Columnmove + IndexChange
    for path in [Left_Genneral_All_path,Right_Genneral_All_path]:
            df1 = pd.read_csv(path, delimiter=',',index_col = 0)
            dfCount = df1.shape
            df1.to_csv(path)
            #write Left to Excel 
            dfValueGeneral = pd.read_csv(path, delimiter=',',usecols = ValueGeneral,nrows= 1)
            dfValueGeneral.to_excel(writer,'General Member',index=False,header=True ,startcol=0,startrow=2)
            if path == Left_Genneral_All_path:
                usecolsArr = LocationOfRowLeft
                LocationMoveColumn = LocationCellForMoveColumn

            else:
                usecolsArr = LocationOfRowRight
                LocationMoveColumn = ExcelCellForMoveColumnRight
            #write Left to Excel 
            """
            DfChangeLeftRight = pd.read_csv(path, delimiter=',',usecols = ValueChangeLeft_Right )
            DfChangeLeftRight.to_excel(writer,'General Member',index=False,header=True ,startcol=0,startrow=20)
            """
            #Write genneral to excel
            DfChangegenneral = pd.read_csv(path, delimiter=',',usecols = GenneralColumnNotChange,nrows= 1 )
            DfChangegenneral.to_excel(writer,'General Member',index=False,header=True ,startcol=0,startrow= int(usecolsArr[1]) )
            #write Genneral Concern Raffter
            DfChangegenneral = pd.read_csv(path, delimiter=',',usecols = GeneralConcernRaffter )
            DfChangegenneral.to_excel(writer,'General Member',index=False,header=True ,startcol=0,startrow=int(usecolsArr[2]))
            #write genneral selected 
            DfChangegenneral = pd.read_csv(path, delimiter=',',usecols = Genneral_Select,nrows= 1)
            DfChangegenneral.to_excel(writer,'General Member',index=False,header=True ,startcol=0,startrow= int(usecolsArr[0]))
            worksheet = writer.sheets['General Member']
            #Wirte Column to csv
            
            for cell,index in zip(LocationMoveColumn,Columnmove):
                cellstr = HandlingDataSTr(cell)
                df = pd.read_csv(path, delimiter=',',index_col = None ,usecols = [index],nrows= 1)
                worksheet.cell(row = cellstr[0], column = cellstr[1]).value = df.iat[0,0]
            IndexChangeTotal = list(ValueChangeLeft_Right)
            GetIndex = GetIndexOfNotChange(IndexChangeRemoveColumn,IndexChangeTotal)
            AligntText(worksheet)
    book.save('new_big_file.xlsx') 
CreateFileExcel()